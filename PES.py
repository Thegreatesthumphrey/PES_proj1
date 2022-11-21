import openpyxl as xl
import xlsxwriter as xw
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill, Border, Color
from openpyxl.styles.differential import DifferentialStyle

# loading
wb = xl.load_workbook("C:\\Users\\KimberlyHumphrey\\WorkProject\\New folder\\rev.xlsx")
ws = wb.active
## trying to maintain pivot tables
pivot = ws._pivots[0] # any will do as they share the same cache
pivot.cache.refreshOnLoad = True
    
# practice
cell_obj = ws.cell(row = 2, column = 1) 
print(cell_obj.value) 


#fill requirements
redFill = PatternFill(start_color='FF0000', end_color='FF0000',
                fill_type='solid')

grnFill = PatternFill(start_color='00FF00', end_color='00FF00',
                fill_type='solid')

yellFill = PatternFill(start_color='FFFF00', end_color='FFFF00',
                fill_type='solid')



#running through payment cells on one sheet (active)
locs = []
values = []

r1 = range(3,11)
r2 = range(2,39)

for x in r1:
    for y in r2:
        cell_obj = ws.cell(y, x)
        if cell_obj.value == None:
            pass
        else:
            locs.append((y,x))
            values.append(cell_obj.value)
            print(cell_obj.value, y, x)
